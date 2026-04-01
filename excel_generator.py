import openpyxl
import shutil
from datetime import datetime

def create_excel(user_id, items):
    template_path = 'template.xlsx'
    output_filename = f"order_{user_id}_{datetime.now().strftime('%d%m_%H%M')}.xlsx"

    # 1. Копируем шаблон, чтобы не испортить оригинал
    shutil.copy(template_path, output_filename)

    # 2. Открываем копию
    wb = openpyxl.load_workbook(output_filename)
    ws = wb.active # Или wb['Название_листа'], если данных много

    # 3. Определяем, с какой строки начинать (допустим, 1-я строка — шапка, пишем со 2-й)
    start_row = 18

    # 4. Записываем данные
    for index, item in enumerate(items):
        current_row = start_row + index
        
        # Записываем строго по колонкам (A=1, B=2, C=3...)
        ws.cell(row=current_row, column=1).value = f'{item['type']} {item['article']} {item['size']} {item['color']}'
        ws.cell(row=current_row, column=2).value = item['brand'] 
        ws.cell(row=current_row, column=3).value = item['gender'] 
        ws.cell(row=current_row, column=4).value = item['article']
        ws.cell(row=current_row, column=5).value = item['color']
        ws.cell(row=current_row, column=6).value = item['size']
        ws.cell(row=current_row, column=7).value = item['count']
        ws.cell(row=current_row, column=8).value = item['compound']
        ws.cell(row=current_row, column=9).value = item['type']    

    # 5. Сохраняем результат
    wb.save(output_filename)
    return output_filename